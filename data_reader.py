import main
import os
import openpyxl

class DataReader:
    def __init__(self, main_program):
        self.main_program = main_program

    def read_dc_folder(self, folder_path):
        """
        Read the dc folder and extract data
        """
        try:
            # Check if the folder path exists
            if not os.path.exists(folder_path):
                raise FileNotFoundError(f"The folder '{folder_path}' does not exist.")

            # Process the files in the folder
            for file_name in os.listdir(folder_path):
                file_path = os.path.join(folder_path, file_name)
                if os.path.isfile(file_path):
                    # Extract data from the file
                    self.extract_data_from_file(file_path)
        except Exception as e:
            print(f"Error occurred while reading the dc folder: {str(e)}")

    def read_mb_excel(self, file_path):
        """
        Read the mb excel file and extract data
        """
        try:
            # Check if the file path exists
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"The file '{file_path}' does not exist.")

            # Load the excel file
            workbook = openpyxl.load_workbook(file_path)

            # Get the active sheet
            sheet = workbook.active

            # Extract data from the sheet
            self.extract_data_from_sheet(sheet)

            # Close the workbook
            workbook.close()
        except Exception as e:
            print(f"Error occurred while reading the mb excel file: {str(e)}")

    def extract_data_from_file(self, file_path):
        """
        Extract data from a file
        """
        try:
            # Process the file based on its extension
            file_extension = os.path.splitext(file_path)[1]
            if file_extension == ".txt":
                # Read the text file and extract data
                with open(file_path, "r") as file:
                    # Process the lines in the file
                    for line in file:
                        # Extract necessary information from the line
                        self.extract_information(line)
            elif file_extension == ".csv":
                # Read the CSV file and extract data
                with open(file_path, "r") as file:
                    # Process the lines in the file
                    for line in file:
                        # Extract necessary information from the line
                        self.extract_information(line.split(","))
            else:
                print(f"Unsupported file format: {file_extension}")
        except Exception as e:
            print(f"Error occurred while extracting data from file '{file_path}': {str(e)}")

    def extract_data_from_sheet(self, sheet):
        """
        Extract data from an excel sheet
        """
        try:
            # Get the maximum row and column count
            max_row = sheet.max_row
            max_column = sheet.max_column

            # Process the rows in the sheet
            for row in range(1, max_row + 1):
                # Extract necessary information from the row
                row_data = []
                for column in range(1, max_column + 1):
                    cell_value = sheet.cell(row=row, column=column).value
                    row_data.append(cell_value)
                self.extract_information(row_data)
        except Exception as e:
            print(f"Error occurred while extracting data from the excel sheet: {str(e)}")

    def extract_information(self, data):
        """
        Extract necessary information from the data
        """
        # Process the data and extract necessary information
        # ...

    def get_person_details(self, person_id):
        """
        Get the details of a person from the data
        """
        try:
            # Retrieve the details of the person based on their ID
            person_details = self.retrieve_person_details(person_id)
            return person_details
        except Exception as e:
            print(f"Error occurred while retrieving person details: {str(e)}")

    def retrieve_person_details(self, person_id):
        """
        Retrieve the details of a person based on their ID
        """
        # Retrieve the details of the person from the data
        # ...

if __name__ == "__main__":
    program = main.MainProgram()
    data_reader = DataReader(program)
    data_reader.read_dc_folder("dc_folder")
    data_reader.read_mb_excel("mb_excel.xlsx")